#!/usr/bin/env python3
import threading
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, scrolledtext, ttk

import openpyxl


class ExcelSheetRenamerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Sheet Renamer")
        self.root.geometry("900x700")
        self.root.resizable(True, True)

        # Color scheme
        self.bg_main = "#f5f5f5"
        self.bg_card = "#ffffff"
        self.primary = "#4f46e5"
        self.primary_hover = "#3730a3"
        self.success = "#10b981"
        self.text_dark = "#1f2937"
        self.text_light = "#6b7280"

        self.root.configure(bg=self.bg_main)
        self.selected_files = []
        self.setup_ui()

    def setup_ui(self):
        # Main container
        main = tk.Frame(self.root, bg=self.bg_main)
        main.pack(fill=tk.BOTH, expand=True, padx=30, pady=30)

        # Title
        title = tk.Label(
            main,
            text="Excel Sheet Renamer",
            font=("Helvetica", 28, "bold"),
            bg=self.bg_main,
            fg=self.text_dark
        )
        title.pack(anchor=tk.W, pady=(0, 5))

        subtitle = tk.Label(
            main,
            text="自动查找订单编号并重命名 Sheet",
            font=("Helvetica", 12),
            bg=self.bg_main,
            fg=self.text_light
        )
        subtitle.pack(anchor=tk.W, pady=(0, 25))

        # Files section
        files_frame = tk.Frame(main, bg=self.bg_card, relief=tk.RAISED, bd=1)
        files_frame.pack(fill=tk.BOTH, expand=False, pady=(0, 15))

        files_inner = tk.Frame(files_frame, bg=self.bg_card)
        files_inner.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)

        # Header
        header_frame = tk.Frame(files_inner, bg=self.bg_card)
        header_frame.pack(fill=tk.X, pady=(0, 10))

        tk.Label(
            header_frame,
            text="文件列表",
            font=("Helvetica", 14, "bold"),
            bg=self.bg_card,
            fg=self.text_dark
        ).pack(side=tk.LEFT)

        self.count_label = tk.Label(
            header_frame,
            text="0 个文件",
            font=("Helvetica", 11),
            bg=self.bg_card,
            fg=self.text_light
        )
        self.count_label.pack(side=tk.LEFT, padx=(10, 0))

        # Listbox
        list_frame = tk.Frame(files_inner, bg=self.bg_card)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.listbox = tk.Listbox(
            list_frame,
            font=("Helvetica", 11),
            height=5,
            bg="#fafafa",
            fg=self.text_dark,
            selectbackground=self.primary,
            selectforeground="white",
            relief=tk.FLAT,
            bd=1,
            yscrollcommand=scrollbar.set
        )
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.listbox.yview)

        # Buttons
        btn_frame = tk.Frame(files_inner, bg=self.bg_card)
        btn_frame.pack(fill=tk.X)

        self.select_btn = tk.Button(
            btn_frame,
            text="选择文件",
            command=self.select_files,
            font=("Helvetica", 12, "bold"),
            bg=self.primary,
            fg="white",
            activebackground=self.primary_hover,
            activeforeground="white",
            relief=tk.RAISED,
            bd=2,
            padx=25,
            pady=10,
            cursor="hand2",
            width=12
        )
        self.select_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.clear_btn = tk.Button(
            btn_frame,
            text="清空",
            command=self.clear_files,
            font=("Helvetica", 12, "bold"),
            bg="#6b7280",
            fg="white",
            activebackground="#4b5563",
            activeforeground="white",
            relief=tk.RAISED,
            bd=2,
            padx=20,
            pady=10,
            cursor="hand2",
            width=8
        )
        self.clear_btn.pack(side=tk.LEFT)

        self.process_btn = tk.Button(
            btn_frame,
            text="开始处理",
            command=self.process_files,
            font=("Helvetica", 12, "bold"),
            bg=self.success,
            fg="white",
            activebackground="#059669",
            activeforeground="white",
            relief=tk.RAISED,
            bd=2,
            padx=25,
            pady=10,
            cursor="hand2",
            state=tk.DISABLED,
            width=12
        )
        self.process_btn.pack(side=tk.RIGHT)

        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress = ttk.Progressbar(
            main,
            variable=self.progress_var,
            maximum=100,
            mode='determinate'
        )
        self.progress.pack(fill=tk.X, pady=(0, 15))

        # Log section
        log_frame = tk.Frame(main, bg=self.bg_card, relief=tk.RAISED, bd=1)
        log_frame.pack(fill=tk.BOTH, expand=True)

        log_inner = tk.Frame(log_frame, bg=self.bg_card)
        log_inner.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)

        tk.Label(
            log_inner,
            text="处理日志",
            font=("Helvetica", 14, "bold"),
            bg=self.bg_card,
            fg=self.text_dark
        ).pack(anchor=tk.W, pady=(0, 10))

        self.log_text = scrolledtext.ScrolledText(
            log_inner,
            font=("Courier", 10),
            bg="#1e293b",
            fg="#e2e8f0",
            relief=tk.FLAT,
            padx=15,
            pady=10,
            state=tk.DISABLED,
            wrap=tk.WORD
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)

    def log(self, msg):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.root.update_idletasks()

    def update_count(self):
        self.count_label.config(text=f"{len(self.selected_files)} 个文件")

    def select_files(self):
        files = filedialog.askopenfilenames(
            title="选择 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )

        if files:
            added = 0
            for f in files:
                if f not in self.selected_files:
                    self.selected_files.append(f)
                    self.listbox.insert(tk.END, Path(f).name)
                    added += 1

            if added > 0:
                self.process_btn.config(state=tk.NORMAL)
                self.update_count()
                self.log(f"✓ 添加了 {added} 个文件")

    def clear_files(self):
        if self.selected_files:
            self.selected_files = []
            self.listbox.delete(0, tk.END)
            self.process_btn.config(state=tk.DISABLED)
            self.update_count()
            self.log("清空文件列表")

    def process_single(self, path):
        try:
            self.log(f"\n{'─' * 60}")
            self.log(f"处理: {Path(path).name}")

            wb = openpyxl.load_workbook(path)
            to_rename = []

            for sheet in wb.worksheets:
                self.log(f"  检查: {sheet.title}")
                order_num = None

                for row in sheet.iter_rows(min_row=1, max_row=20, min_col=1, max_col=10):
                    for cell in row:
                        if cell.value and "订单编号" in str(cell.value) and cell.row and cell.column:
                            below = sheet.cell(row=cell.row + 1, column=cell.column)
                            if below.value:
                                order_num = str(below.value).strip()
                                self.log(f"    → 找到: {order_num}")
                            break
                    if order_num:
                        break

                if order_num:
                    to_rename.append((sheet.title, order_num))
                else:
                    self.log("    未找到订单编号")

            renamed = 0
            for old, new in to_rename:
                try:
                    sheet = wb[old]
                    if new in wb.sheetnames and new != old:
                        self.log(f"    跳过: {new} 已存在")
                        continue
                    sheet.title = new
                    self.log(f"    ✓ {old} → {new}")
                    renamed += 1
                except Exception as e:
                    self.log(f"    ✗ 失败: {e}")

            out = path.replace('.xlsx', '_renamed.xlsx')
            wb.save(out)
            self.log(f"✓ 保存: {Path(out).name} (重命名 {renamed} 个)")

            # 拆分 sheet 为单个文件
            self.log("\n开始拆分 sheet...")
            split_count = self.split_sheets(wb, path)
            self.log(f"✓ 拆分完成: 生成 {split_count} 个文件")

            return True

        except Exception as e:
            self.log(f"✗ 错误: {e}")
            return False

    def split_sheets(self, wb, original_path):
        """将每个 sheet 拆分成单个 excel 文件"""
        # 创建输出目录
        path_obj = Path(original_path)
        base_name = path_obj.stem  # 不包含扩展名的文件名
        output_dir = path_obj.parent / base_name

        # 如果目录不存在则创建
        if not output_dir.exists():
            output_dir.mkdir(parents=True)
            self.log(f"  创建目录: {output_dir.name}")

        split_count = 0

        # 获取所有 sheet 名称
        sheet_names = [sheet.title for sheet in wb.worksheets]

        # 对每个 sheet 创建单独的文件
        for sheet_name in sheet_names:
            try:
                # 重新加载原始文件（保留所有原始格式）
                new_wb = openpyxl.load_workbook(
                    original_path.replace('.xlsx', '_renamed.xlsx')
                )

                # 删除除当前 sheet 外的所有 sheet
                sheets_to_remove = [s for s in new_wb.sheetnames if s != sheet_name]
                for s in sheets_to_remove:
                    new_wb.remove(new_wb[s])

                # 使用 sheet 名称作为文件名
                file_name = f"{sheet_name}.xlsx"
                output_path = output_dir / file_name

                # 保存文件
                new_wb.save(output_path)
                self.log(f"  ✓ 生成: {file_name}")
                split_count += 1

            except Exception as e:
                self.log(f"  ✗ 拆分 {sheet_name} 失败: {e}")

        return split_count

    def process_thread(self):
        total = len(self.selected_files)
        success = 0

        self.log(f"\n{'═' * 60}")
        self.log(f"开始处理 {total} 个文件")
        self.log(f"{'═' * 60}")

        for idx, path in enumerate(self.selected_files, 1):
            if self.process_single(path):
                success += 1

            self.progress_var.set((idx / total) * 100)
            self.root.update_idletasks()

        self.log(f"\n{'═' * 60}")
        self.log(f"完成! 成功: {success}/{total}")
        self.log(f"{'═' * 60}")

        messagebox.showinfo("完成", f"处理完成!\n\n成功: {success}/{total}")

        self.select_btn.config(state=tk.NORMAL)
        self.process_btn.config(state=tk.NORMAL)
        self.clear_btn.config(state=tk.NORMAL)

    def process_files(self):
        if not self.selected_files:
            messagebox.showwarning("提示", "请先选择文件")
            return

        self.select_btn.config(state=tk.DISABLED)
        self.process_btn.config(state=tk.DISABLED)
        self.clear_btn.config(state=tk.DISABLED)
        self.progress_var.set(0)

        threading.Thread(target=self.process_thread, daemon=True).start()


def main():
    root = tk.Tk()
    app = ExcelSheetRenamerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
